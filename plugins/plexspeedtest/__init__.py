import threading
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, List, Dict, Tuple

import pytz
from apscheduler.schedulers.background import BackgroundScheduler
from openpyxl import load_workbook

from app.core.config import settings
from app.core.plugin import PluginManager
from app.log import logger
from app.plugins import _PluginBase
from app.scheduler import Scheduler

lock = threading.Lock()


class PlexSpeedTest(_PluginBase):
    # 插件名称
    plugin_name = "Plex IP优选"
    # 插件描述
    plugin_desc = "自动获取Plex相关域名，实现IP优选。"
    # 插件图标
    plugin_icon = "https://raw.githubusercontent.com/InfinityPacer/MoviePilot-Plugins/main/icons/plexspeedtest.png"
    # 插件版本
    plugin_version = "1.0"
    # 插件作者
    plugin_author = "InfinityPacer"
    # 作者主页
    author_url = "https://github.com/InfinityPacer"
    # 插件配置项ID前缀
    plugin_config_prefix = "PlexSpeedTest_"
    # 加载顺序
    plugin_order = 97
    # 可使用的用户级别
    auth_level = 1

    # region 私有属性

    pluginmanager = None

    # 是否开启
    _enabled = False
    # 立即运行一次
    _onlyonce = False
    # 定时器
    _scheduler = None
    # 退出事件
    _event = threading.Event()

    # endregion

    def init_plugin(self, config: dict = None):
        self.pluginmanager = PluginManager()

        if not config:
            return

        self._enabled = config.get("enabled")
        self._onlyonce = config.get("onlyonce")

        if not self._enabled:
            self.action()

        if self._enabled and self._onlyonce:
            # 启动服务
            self._scheduler = BackgroundScheduler(timezone=settings.TZ)
            self._scheduler.add_job(
                func=self.action,
                trigger="date",
                run_date=datetime.now(tz=pytz.timezone(settings.TZ)) + timedelta(seconds=3),
                name=f"{self.plugin_name}",
            )
            self._onlyonce = False
            config["onlyonce"] = False
            self.update_config(config=config)

            # 启动服务
            if self._scheduler.get_jobs():
                self._scheduler.print_jobs()
                self._scheduler.start()

    def get_state(self) -> bool:
        return self._enabled

    @staticmethod
    def get_command() -> List[Dict[str, Any]]:
        """
        定义远程控制命令
        :return: 命令关键字、事件、描述、附带数据
        """
        pass

    def get_api(self) -> List[Dict[str, Any]]:
        pass

    def get_form(self) -> Tuple[List[dict], Dict[str, Any]]:
        """
        拼装插件配置页面，需要返回两块数据：1、页面配置；2、数据结构
        """
        return [
            {
                'component': 'VForm',
                'content': [
                    {
                        'component': 'VRow',
                        'content': [
                            {
                                'component': 'VCol',
                                'props': {
                                    'cols': 12,
                                    'md': 4
                                },
                                'content': [
                                    {
                                        'component': 'VSwitch',
                                        'props': {
                                            'model': 'enabled',
                                            'label': '启用插件',
                                            'hint': '开启后插件将处于激活状态',
                                            'persistent-hint': True
                                        },
                                    }
                                ]
                            },
                            {
                                'component': 'VCol',
                                'props': {
                                    'cols': 12,
                                    'md': 4
                                },
                                'content': [
                                    {
                                        'component': 'VSwitch',
                                        'props': {
                                            'model': 'onlyonce',
                                            'label': '立即运行一次',
                                            'hint': '插件将立即运行一次',
                                            'persistent-hint': True
                                        }
                                    }
                                ]
                            }
                        ]
                    },
                    {
                        'component': 'VRow',
                        'content': [
                            {
                                'component': 'VCol',
                                'props': {
                                    'cols': 12,
                                },
                                'content': [
                                    {
                                        'component': 'VAlert',
                                        'props': {
                                            'type': 'info',
                                            'variant': 'tonal',
                                            'text': '注意：本插件依赖自定义Hosts以及Cloudflare IP优选插件，请提前安装对应插件并进行相关配置'
                                        }
                                    }
                                ]
                            }
                        ]
                    },
                    {
                        'component': 'VRow',
                        'content': [
                            {
                                'component': 'VCol',
                                'props': {
                                    'cols': 12,
                                },
                                'content': [
                                    {
                                        'component': 'VAlert',
                                        'props': {
                                            'type': 'info',
                                            'variant': 'tonal',
                                            'text': '注意：请提前在Cloudflare IP优选插件中配置优选IP，若无，可配置为 121.121.121.121'
                                        }
                                    }
                                ]
                            }
                        ]
                    },
                    {
                        'component': 'VRow',
                        'content': [
                            {
                                'component': 'VCol',
                                'props': {
                                    'cols': 12,
                                },
                                'content': [
                                    {
                                        'component': 'VAlert',
                                        'props': {
                                            'type': 'info',
                                            'variant': 'tonal',
                                            'text': '注意：本插件启用后，将启用自定义Hosts插件并写入相关信息，同时将运行一次Cloudflare IP优选插件'
                                        }
                                    }
                                ]
                            }
                        ]
                    }
                ]
            }
        ], {
            "enabled": False,
            "onlyonce": False
        }

    def get_page(self) -> List[dict]:
        pass

    def get_service(self) -> List[Dict[str, Any]]:
        """
        注册插件公共服务
        [{
            "id": "服务ID",
            "name": "服务名称",
            "trigger": "触发器：cron/interval/date/CronTrigger.from_crontab()",
            "func": self.xxx,
            "kwargs": {} # 定时器参数
        }]
        """
        pass

    def stop_service(self):
        """
        退出插件
        """
        pass

    def action(self):
        """
        执行服务
        """
        with lock:
            logger.info(f"开始执行{self.plugin_name}服务")

            result, message = self.__check_required_plugins_installed()
            if not result:
                self.__log_and_notify(message=message)
                return

            if self._enabled:
                dns_records = self.__get_dns_records()
                cf_ip = self.__get_cf_ip()
                if self.__write_hosts(dns_records=dns_records, cf_ip=cf_ip):
                    if cf_ip:
                        self.__speed_test()
            else:
                self.__restore_hosts()

    def __speed_test(self):
        """
        执行测速
        """
        logger.info("正在准备运行一次CloudflareSpeedTest")
        config = self.get_config(plugin_id="CloudflareSpeedTest") or {}
        config["onlyonce"] = True
        self.update_config(config=config, plugin_id="CloudflareSpeedTest")
        self.__reload_plugin(plugin_id="CloudflareSpeedTest")
        logger.info("已通知CloudflareSpeedTest运行一次")

    def __write_hosts(self, dns_records: List[dict], cf_ip: str) -> bool:
        """
        写入Hosts
        """
        logger.info("正在准备写入Hosts")
        if not dns_records:
            logger.warn("Plex DNS 记录为空，跳过写入")
            return False

        logger.info(f"Plex DNS 记录: {dns_records}")

        config = self.get_config(plugin_id="CustomHosts") or {}
        config["enabled"] = True
        hosts_str = config.get("hosts", "")

        start_marker = "# PlexSpeedTest Begin"
        end_marker = "# PlexSpeedTest End"
        new_plex_hosts = "\n".join(
            [f"{cf_ip if cf_ip else record['IP Address']} {record['Hostname']}" for record in dns_records])
        new_hosts_str = self.__update_hosts_content(hosts_str, new_plex_hosts, start_marker, end_marker)

        config["hosts"] = new_hosts_str
        self.update_config(config=config, plugin_id="CustomHosts")
        self.__reload_plugin(plugin_id="CustomHosts")
        logger.info("Hosts写入完成")
        return True

    @staticmethod
    def __update_hosts_content(hosts_str, new_block, start_marker, end_marker):
        """
        更新Hosts文件内容
        """
        start_index = hosts_str.find(start_marker)
        end_index = hosts_str.find(end_marker, start_index)
        if start_index != -1 and end_index != -1:
            return f"{hosts_str[:start_index + len(start_marker)]}\n{new_block}\n{hosts_str[end_index:]}"
        else:
            return f"{hosts_str}\n{start_marker}\n{new_block}\n{end_marker}"

    def __restore_hosts(self):
        """
        还原Hosts文件
        """
        logger.info("正在准备还原Hosts")
        config = self.get_config(plugin_id="CustomHosts") or {}
        hosts_str = config.get("hosts", "")
        if not hosts_str:
            logger.info("Hosts配置为空，跳过还原")
            return

        start_marker = "# PlexSpeedTest Begin"
        end_marker = "# PlexSpeedTest End"
        new_hosts_str = self.__remove_hosts_block(hosts_str, start_marker, end_marker)

        if new_hosts_str != hosts_str:  # 只有在内容变更时才更新配置
            config["hosts"] = new_hosts_str
            self.update_config(config=config, plugin_id="CustomHosts")
            self.__reload_plugin(plugin_id="CustomHosts")
            logger.info("Hosts还原完成")
        else:
            logger.info("没有找到特定块，无需还原Hosts")

    @staticmethod
    def __remove_hosts_block(hosts_str, start_marker, end_marker):
        """
        从Hosts字符串中移除指定的块
        """
        start_index = hosts_str.find(start_marker)
        end_index = hosts_str.find(end_marker, start_index)
        if start_index != -1 and end_index != -1:
            # 确保移除后不会留下多余的空行
            before_block = hosts_str[:start_index].rstrip()
            after_block = hosts_str[end_index + len(end_marker):].lstrip()
            return f"{before_block}\n{after_block}" if before_block and after_block else before_block or after_block
        return hosts_str

    def __get_cf_ip(self) -> str:
        """
        获取CloudFlare Ip
        """
        cf_ip = ""
        cf_config = self.get_config(plugin_id="CloudflareSpeedTest")
        if cf_config:
            cf_ip = cf_config.get("cf_ip")
        if cf_ip:
            logger.info(f"从CloudflareSpeedTest配置中获取到IP: {cf_ip}")
        else:
            logger.info("CloudflareSpeedTest配置中没有找到IP，将使用DNS记录中的IP")
        return cf_ip

    def __get_dns_records(self) -> List[dict]:
        """
        获取Plex DNS 记录
        """
        default_records = self.__get_default_records()
        try:
            filepath = Path(__file__).parent / "plex.tv_dns.xlsx"
            sheet = self.__load_excel(filepath)
            result_records = {}
            if sheet is not None:
                # 获取表头与列索引的映射关系
                headers = self.__get_column_indices(sheet)
                if headers:
                    result_records = self.__filter_records(sheet, headers)
            if not result_records:
                logger.warn(f"没有获取在线 Plex DNS 记录，使用默认 DNS 记录")
                result_records = default_records
            return result_records
        except Exception as e:
            logger.error(f"获取Plex DNS 记录发生异常，使用默认 DNS 记录，{e}")
            return default_records

    @staticmethod
    def __get_default_records() -> List[dict]:
        """
        获取默认记录
        """
        return [
            {"Hostname": "app-qa.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "meta.plex.tv", "IP Address": "104.18.27.211"},
            {"Hostname": "chapterdb.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "static.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "metadata-static.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "website-static.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "provider-static.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "bundle-archive-codeload.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "acoustid.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "lyricfind.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "analytics-stage.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "chapterdb-archive.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "bundle-archive.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "watch-staging.plex.tv", "IP Address": "104.18.27.211"},
            {"Hostname": "images-staging.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "sonos-staging.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "gist-staging.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "blog.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "watch.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "songkick.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "email.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "mpm.plex.tv", "IP Address": "104.18.26.211"},
            {"Hostname": "help.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "plexamp.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "ump.plex.tv", "IP Address": "104.18.26.211"},
            {"Hostname": "app.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "metadata.provider.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "music.provider.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "vod.provider.plex.tv", "IP Address": "104.18.27.211"},
            {"Hostname": "music-staging.provider.plex.tv", "IP Address": "104.18.26.211"},
            {"Hostname": "epg-staging.provider.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "email-staging.provider.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "discover-staging.provider.plex.tv", "IP Address": "104.18.27.211"},
            {"Hostname": "play-staging.provider.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "epg.provider.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "email.provider.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "discover.provider.plex.tv", "IP Address": "104.18.26.211"},
            {"Hostname": "vod2-dev.provider.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "legadata-dev.provider.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "metadata-dev.provider.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "music-dev.provider.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "vod-dev.provider.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "datadog-dev.provider.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "epg-dev.provider.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "discover-dev.provider.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "news-dev.provider.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "play-dev.provider.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "play.provider.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "transcoder.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "cordcutter.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "staging1-cordcutter.plex.tv", "IP Address": "104.18.27.211"},
            {"Hostname": "staging3-cordcutter.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "ump-paas.plex.tv", "IP Address": "104.18.26.211"},
            {"Hostname": "analytics.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "downloads.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "images.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "nightlies.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "scrobbles.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "links.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "plugins.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "sonos.plex.tv", "IP Address": "104.18.27.211"},
            {"Hostname": "htbackdrops.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "artifacts.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "assets.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "clients.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "support.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "staging1-support.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "staging3-support.plex.tv", "IP Address": "104.18.27.211"},
            {"Hostname": "gist.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "images-dev.plex.tv", "IP Address": "172.64.146.103"},
            {"Hostname": "sonos-dev.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "www.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "staging1-www.plex.tv", "IP Address": "104.18.27.211"},
            {"Hostname": "staging3-www.plex.tv", "IP Address": "104.18.41.153"},
            {"Hostname": "mbz.plex.tv", "IP Address": "172.64.146.103"}
        ]

    @staticmethod
    def __load_excel(file_path):
        """
        读取 Excel 文件并返回工作表的对象
        :param file_path: Excel 文件路径
        :return: 工作表对象
        """
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            return sheet
        except FileNotFoundError:
            logger.error(f"The file '{file_path}' was not found.")
            return None
        except Exception as e:
            logger.error(f"An unexpected error occurred while loading the file: {e}")
            return None

    @staticmethod
    def __get_column_indices(sheet):
        """
        获取表头中每一列对应的索引位置
        :param sheet: 工作表对象
        :return: 列名与列索引的字典
        """
        headers = {}
        try:
            for idx, cell in enumerate(sheet[1], 1):  # 表头通常在第一行
                if cell.value:
                    headers[cell.value] = idx - 1  # 将表头名与列索引映射，idx从1开始，转为0基
            return headers
        except Exception as e:
            logger.error(f"An error occurred while reading the header: {e}")
            return headers

    @staticmethod
    def __filter_records(sheet, headers):
        """
        筛选 'Type' 为 'A' 且 'Netblock Owner' 包含 'CloudFlare' 的记录
        :param sheet: 工作表对象
        :param headers: 表头与列索引的映射字典
        :return: 筛选后的记录列表
        """
        records = []
        try:
            type_idx = headers.get("Type")
            netblock_owner_idx = headers.get("Netblock Owner")
            hostname_idx = headers.get("Hostname")
            ip_address_idx = headers.get("IP Address")

            for row in sheet.iter_rows(min_row=2, values_only=True):  # 跳过第一行（表头）
                type_value = row[type_idx] if type_idx is not None else None
                netblock_owner = row[netblock_owner_idx] if netblock_owner_idx is not None else None
                if type_value and netblock_owner:
                    if type_value.upper() == "A" and "CloudFlare".lower() in netblock_owner.lower():
                        record = {
                            "Hostname": row[hostname_idx] if hostname_idx is not None else None,
                            "IP Address": row[ip_address_idx] if ip_address_idx is not None else None
                        }
                        records.append(record)
        except Exception as e:
            logger.error(f"An error occurred while filtering records: {e}")

        return records

    def __check_required_plugins_installed(self) -> (bool, str):
        """
        检查所有指定的依赖插件是否已安装
        """
        plugin_names = {
            "CustomHosts": "自定义Hosts",
            "CloudflareSpeedTest": "Cloudflare IP优选"
        }

        # 获取本地插件列表
        local_plugins = self.pluginmanager.get_local_plugins()

        # 初始化未安装插件列表
        missing_plugins = []

        # 校验所有的插件是否已安装
        for plugin_id, plugin_name in plugin_names.items():
            plugin = next((p for p in local_plugins if p.id == plugin_id and p.installed), None)
            if not plugin:
                missing_plugins.append(plugin_name)

        if missing_plugins:
            missing_plugins_str = "，".join(missing_plugins)
            return False, f"以下插件尚未安装: {missing_plugins_str}"

        return True, "所有指定插件均已安装"

    def __log_and_notify(self, message):
        """
        记录日志并发送系统通知
        """
        logger.info(message)
        self.systemmessage.put(message, title=f"{self.plugin_name}")

    def __reload_plugin(self, plugin_id: str):
        """
        热加载
        """
        logger.info(f"准备热加载插件: {plugin_id}")

        # 加载插件到内存
        try:
            self.pluginmanager.reload_plugin(plugin_id)
            logger.info(f"成功热加载插件: {plugin_id} 到内存")
        except Exception as e:
            logger.error(f"失败热加载插件: {plugin_id} 到内存. 错误信息: {e}")
            return

        # 注册插件服务
        try:
            Scheduler().update_plugin_job(plugin_id)
            logger.info(f"成功热加载插件到插件服务: {plugin_id}")
        except Exception as e:
            logger.error(f"失败热加载插件到插件服务: {plugin_id}. 错误信息: {e}")
            return

        logger.info(f"已完成插件热加载: {plugin_id}")
